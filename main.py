import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api import VkUpload
import hashlib
import openpyxl
from openpyxl import load_workbook
import random
import pandas as pd

vk_session = vk_api.VkApi(token = 'vk1.a.wYAgJO6ytwv67R1CB2UXuQCHqvv-3RDhK54ngq8RZ5Dp-rk3C49zVY0o5-0lJCdwWIyf0PlzoCHoL5YEXT-n4ejvy3V_6-QTuPGM5mSdpGzQ-ZEEu0k-KlRklF-L4K0moK7mWEtbxJe9jg83yKYoy847ZMd9qUmEjRR47x2vsvKhG-oHKlG-2OFOVmwLzkOpuRngfyFjroCeOShXr3nJGw')
session_api = vk_session.get_api()
longpoll = VkLongPoll(vk_session)
upload = VkUpload(session_api)

def send_msg(id, text, keyboard=None):
    post = {
        "user_id": id,
        "message": text,
        "random_id": 0
    }
    if keyboard != None:
        post["keyboard"] = keyboard.get_keyboard()
    vk_session.method("messages.send", post)
    return True


def get_user_info(user_id):
    user_info = session_api.users.get(user_ids=user_id, fields='first_name,last_name')
    first_name = user_info[0]['first_name']
    last_name = user_info[0]['last_name']
    return (first_name + " " + last_name)


def sheets(id, score, file_path = 'users_data.xlsx'):
    wb = load_workbook(file_path)
    sheet = wb.active
    row_index = None
    for i, row in enumerate(sheet.rows):
        if row[0].value == id:
            row_index = i + 1
            break
    if row_index is None:
        row_index = len(sheet['A']) + 1
        sheet.cell(row=row_index, column=1).value = id
        sheet.cell(row=row_index, column=2).value = score
    else:
        current_score = sheet.cell(row=row_index, column=2).value
        new_score = current_score or 0
        new_score += score
        sheet.cell(row=row_index, column=2).value = new_score
    wb.save(file_path)

def code_check(msg, id):
    checklist = open('checklist.txt', "r").readlines()
    keyboard = VkKeyboard()
    keyboard.add_button("В меню", VkKeyboardColor.SECONDARY)
    if msg in checklist:
        sheets(get_user_info(id), score)
        send_msg(id, "Ваше участие зарегистрировано!")
        return True
    else:
        send_msg(id, "Такого кода мероприятия не существует, либо он введен неправильно. "
                     "Проверьте правильность ввода и повторите попытку")
        return False


def result():
    df = pd.read_excel("users_data.xlsx")
    if len(df) < 10:
        selected_rows = df
    else:
        selected_rows = df.head(10)
    selected_rows.to_excel("result.xlsx", index=False)

def code_generate(id, score):
    checklist = open('checklist.txt', "r+")
    chars = 'abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
    password = ''
    for i in range(6):
        password += random.choice(chars)

    keyboard = VkKeyboard()
    keyboard.add_button("В меню", VkKeyboardColor.SECONDARY)
    send_msg(id, "Готово! Ваш код сгенерирован", keyboard)
    send_msg(id, password, keyboard)
    checklist.write(password)

    return True


score = 100
con = False
#adm = open("admin_data.txt").readlines()
adm = "353608831"

for event in longpoll.listen():
    if event.type == VkEventType.MESSAGE_NEW:
        if event.to_me:
            msg = event.text.lower()
            id = event.user_id

            if msg == "панель администратора" and str(id) in adm:
                keyboard = VkKeyboard()
                keyboard.add_button("Сгенерировать код", VkKeyboardColor.PRIMARY)
                keyboard.add_button("Текущий рейтинг", VkKeyboardColor.PRIMARY)
                send_msg(id, "Добро пожаловать! Что сделаем сегодня?", keyboard)

            if msg == "сгенерировать код":
                code_generate(id, score)

            if msg == "текущий рейтинг" and str(id) in adm:
                photo = upload.document_message(doc= 'result.xlsx', peer_id = id)
                # Отправка сообщения с файлом пользователю
                session_api.messages.send(
                    user_id = id,
                    message = 'Рейтинг сформирован',
                    attachment = f"doc{photo['doc']['owner_id']}_{photo['doc']['id']}",
                    random_id = 0
                )


            if msg == "начать":
                keyboard = VkKeyboard()
                keyboard.add_button("Ввести код", VkKeyboardColor.PRIMARY)
                send_msg(id, "Что вы хотите сделать?", keyboard)

            if msg == "ввести код":
                send_msg(id, "Укажите уникальный идентификатор мероприятия", None)
                con = True

            if con == True and len(msg) == 6:
                code_check(msg, id)

            if msg == "в меню":
                keyboard = VkKeyboard()
                keyboard.add_button("Начать", VkKeyboardColor.PRIMARY)
                keyboard.add_button("Панель администратора", VkKeyboardColor.SECONDARY)
                send_msg(id, "Добро пожаловать! Что сделаем сегодня?", keyboard)


            else:
                if msg == None:
                    keyboard = VkKeyboard()
                    keyboard.add_button("В меню", VkKeyboardColor.PRIMARY)
                    send_msg(id, "Я вас не понимаю... (незвестная команда)", keyboard)
